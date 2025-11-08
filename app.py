import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_file(filename):
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']

        for i in range(2,sheet.max_row+1):
            cell= sheet.cell(row=i, column=3)
            corrected_value = cell.value * 0.9
            corrected_value_cell = sheet.cell(i,4 )
            corrected_value_cell.value= corrected_value

        values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart)

        if filename.lower().endswith('.xlsx'):
            new_filename = filename[:-5] + '_new.xlsx'
        else:
            new_filename = filename + '_new.xlsx'

        wb.save(new_filename)
        print(f"Saved new file as: {new_filename}")


process_file('transactions.xlsx')